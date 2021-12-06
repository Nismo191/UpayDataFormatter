# Import Standard Libraries
import sys
sys.path.append("..")

# Import Additional Libraries
import pandas as pd
from openpyxl import load_workbook

# Import Custom Libraries
from BenLogic import file_operations as fo
import key_data as data

# Initialize Variables
sites_array = data.sites_array
units_array = data.units_array
card_rates = data.card_rates
site = ""
unit = ""
rate = "0"
sum_transaction_value = 0
sum_transaction_charge = 0
sum_standard_charge = 0
sum_service_charge = 0
df_temp = []
totals = []

# Load data and template file, relavent data starts at row 48
def load_files(data_url, template_url):
    data = fo.read_excel(data_url, 48)
    template = load_workbook(template_url)
    return data, template

# Unit class holds all of the calculations for a single unit
class Unit():
    def __init__(self, site, unit, scheme, tVal, tCharge, standCharge, serviceCharge):
        self.site = site
        self.unit = unit
        self.scheme = scheme
        self.sum_transaction_value = tVal
        self.sum_transaction_charge = tCharge
        self.sum_standard_charge = standCharge
        self.sum_service_charge = serviceCharge

    def __str__(self):
        return f"{self.site} | {self.unit} | {self.scheme} | {self.sum_transaction_value} | {self.sum_transaction_charge} | {self.sum_standard_charge} | {self.sum_service_charge}"
    


if __name__ == "__main__":
    # Load data and template file, data is returned as a dataframe, template is returned as a openpyxl workbook
    data, template = load_files('Input/Upay example- 1.xlsx', 'Input/1800 Upay Journal - Copy.xlsx')

    # Remove unneeded columns (anything called 'Unnamed')
    for col in data.columns:
        if col.startswith('Unnamed'):
            data.drop(col, axis=1, inplace=True)

    # Loop through each row in the dataframe and set the site/unit var each time the site/unit changes and pick up the correct rate based on card type.
    # Append the site/unit/rate to the current row and then append the row to the df_temp array.
    for row in data.itertuples():
        # if row.Type contains 'Norton Canes'
        if any(x in str(row.Type) for x in sites_array):
            site = row.Type[5:]
        if any(x in str(row.Type) for x in units_array):
            unit = row.Type[6:]
        # Check if key is in card_rates
        if row._2 in card_rates:
            rate = card_rates[row._2]['rate']
        row = row + (site,unit,float(row.Count) * float(rate))
        df_temp.append(row)

    # Build a new pandas dataframe from the df_temp array
    df = pd.DataFrame(df_temp)

    # Loop through rows and drop all site/unit rows
    for row in df.itertuples():
        if pd.isna(row._3):
            df.drop(row.Index, axis=0, inplace=True)

    # Set the dataframe column names back to the original names, they are removed when the rows are written to the temp array above
    df.columns = ['Index', 'Type', 'Scheme', 'TCount', 'TVal', 'Standard Charge Rate', 'Standard Charge Value', 'Service Charge Rate', 'Service Charge Value', 'Site', 'Unit', 'Service Charge Value Corrected']

    # Output the dataframe to a csv file for debugging
    # DataFrame.to_csv(df, 'output/output.csv', index=False)

    # Loop through all Site/unit combinations. Using DataFrame.query() get the rows for the current combo and sum relavent columns 
    # Append a new Unit object containing the set variables to the totals array
    # Reset the variables for the next combo
    for site in sites_array:
        for unit in units_array:
            sum_transaction_value = 0
            sum_transaction_charge = 0
            sum_standard_charge = 0
            sum_service_charge = 0
            if df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme != "American Express"')['TVal'].sum() != 0:
                sum_transaction_value += df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme != "American Express"')['TVal'].sum()
                sum_transaction_charge += (df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme != "American Express"')['Service Charge Value Corrected'].sum())
                sum_transaction_charge += df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme != "American Express"')['Standard Charge Value'].sum()
                sum_standard_charge += df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme != "American Express"')['Service Charge Value'].sum()
                sum_service_charge = (sum_standard_charge/100)*20
                sum_transaction_value = float("{:.2f}".format(sum_transaction_value*-1))
                sum_transaction_charge = float("{:.2f}".format(sum_transaction_charge))
                sum_standard_charge = float("{:.2f}".format(sum_standard_charge))
                sum_service_charge = float("{:.2f}".format(sum_service_charge))
                totals.append(Unit(site, unit, "", sum_transaction_value, sum_transaction_charge, sum_standard_charge, sum_service_charge))
            # Reset the sum variable for the amex record
            sum_transaction_value = 0
            sum_transaction_charge = 0
            sum_standard_charge = 0
            sum_service_charge = 0
            if df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme == "American Express"')['TVal'].sum() != 0:
                sum_transaction_value += df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme == "American Express"')['TVal'].sum()
                sum_transaction_charge += (df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme == "American Express"')['Service Charge Value Corrected'].sum())
                sum_transaction_charge += (df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme == "American Express"')['Standard Charge Value'].sum())
                sum_standard_charge += df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme == "American Express"')['Service Charge Value'].sum()
                sum_service_charge = (sum_standard_charge/100)*20
                sum_transaction_value = float("{:.2f}".format(sum_transaction_value*-1))
                sum_transaction_charge = float("{:.2f}".format(sum_transaction_charge))
                sum_standard_charge = float("{:.2f}".format(sum_standard_charge))
                sum_service_charge = float("{:.2f}".format(sum_service_charge))
                totals.append(Unit(site, unit, "Amex", sum_transaction_value, sum_transaction_charge, sum_standard_charge, sum_service_charge))

    # Set the active sheet for the template wookbook and get the read and write ranges
    ws = template.active
    read_range = ws['F10':'F146']
    write_range = ws['C10':'C146']
    total_range = ws['C10']

    #  Go through each record and amend site/unit names to match the template
    ###### Should switch this to map an ID to the site/unit name and then use that to get the correct row in the template ######
    for record in totals:
        if 'Costa' in record.unit:
            record.unit = record.unit.replace('Coffee', '').strip()
        if 'WHS' in record.unit:
            record.unit = record.unit.replace('WHSmiths', 'WHSmith').strip()
        if 'North' in record.site:
            record.site = record.site.replace('(North)', 'Nth').strip()
        if 'South' in record.site:
            record.site = record.site.replace('(South)', 'Sth').strip()
            
    # Print all records to the console for debugging
    for record in totals:
        print(record.__str__())

    # Loop through the read range and then the totals array, where the Site/Unit/Scheme match, write the values to the write range for the current index
    # increment the index for the three additional vars
    i=0
    total = 0
    for cell in read_range:
        for record in totals:
            # print(cell[0].value, record.site + " " + record.unit)
            if cell[0].value == (record.site + " " + record.unit + " " + record.scheme).strip(): 
                write_range[i][0].value = record.sum_transaction_value
                write_range[i+1][0].value = record.sum_transaction_charge
                write_range[i+2][0].value = record.sum_standard_charge
                write_range[i+3][0].value = record.sum_service_charge
                if record.scheme != "Amex":
                    total += record.sum_transaction_value
        i += 1

    total_range.value = total

    # Write the template to the output file
    template.save('output/output.xlsx')